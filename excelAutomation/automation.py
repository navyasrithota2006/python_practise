import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb.active
    sheet.cell(1,11).value = "Discounted Value"

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row,6)
        if isinstance(cell.value, (int, float)):
            corrected_value = cell.value * 0.9
        else:
            corrected_value = 0
        sheet.cell(row,11).value = corrected_value

    values = Reference(sheet,
                    min_row=2,
                    max_row=sheet.max_row,
                    min_col = 11,
                    max_col = 11)

    chart = BarChart()
    chart.add_data(values)
    chart.title = "Discounted Values"
    sheet.add_chart(chart, 'm2')
    wb.save(filename)

process_workbook('Book1.xlsx')
